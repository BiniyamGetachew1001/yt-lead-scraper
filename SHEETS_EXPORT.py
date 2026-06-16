#!/usr/bin/env python3
"""
SHEETS_EXPORT.py — Excel CRM Export
Converts the scraped lead list into a formatted, outreach-ready Excel workbook.

Zero cloud accounts required. No API keys. Completely free.
Output: leads_export.xlsx  (opens in Excel, Google Sheets, LibreOffice)
"""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

SHEET_HEADERS: List[str] = [
    "Channel Name",
    "Channel URL",
    "Subscribers",
    "Total Views",
    "Upload Age (Days)",
    "Email",
    "Instagram",
    "TikTok",
    "Twitter / X",
    "LinkedIn",
    "Other Links",
    "Channel Description",
    "Keyword Found By",
    "Outreach Status",
    "Notes",
    "Date Added",
]

# Maps lead dict key → 1-based column index (must match SHEET_HEADERS order)
_FIELD_TO_COL: Dict[str, int] = {
    "channel_name":       1,
    "channel_url":        2,
    "subscribers":        3,
    "total_views":        4,
    "days_since_upload":  5,
    "email":              6,
    "instagram":          7,
    "tiktok":             8,
    "twitter":            9,
    "linkedin":           10,
    "other_links":        11,
    "description":        12,
    "keyword_found_by":   13,
    "outreach_status":    14,
    # col 15 = Notes  → blank for user to fill
    "scraped_at":         16,
}

# Approximate column widths in Excel units
_COL_WIDTHS: Dict[int, int] = {
    1: 28,   # Channel Name
    2: 42,   # Channel URL
    3: 14,   # Subscribers
    4: 14,   # Total Views
    5: 18,   # Upload Age
    6: 30,   # Email
    7: 30,   # Instagram
    8: 30,   # TikTok
    9: 30,   # Twitter
    10: 30,  # LinkedIn
    11: 40,  # Other Links
    12: 50,  # Description
    13: 22,  # Keyword
    14: 18,  # Outreach Status
    15: 30,  # Notes
    16: 20,  # Date Added
}

# Outreach status dropdown options
_STATUS_OPTIONS = [
    "Not Contacted",
    "Email Sent",
    "Followed Up",
    "Replied",
    "Call Booked",
    "Converted",
    "Not Interested",
    "Bounced",
]

# Colour palette
_HEADER_BG   = "1A1D2B"   # dark navy (hex, no #)
_HEADER_FG   = "FFFFFF"   # white text
_ALT_ROW_BG  = "F3F4F8"   # very light grey for alternating rows
_ACCENT_BLUE = "4299E1"   # URL column tint
_GREEN_BG    = "E6F9F0"   # rows with email found
_LINK_BLUE   = "2B6CB0"   # hyperlink colour


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _header_font()   -> Font:
    return Font(name="Calibri", bold=True, color=_HEADER_FG, size=10)

def _header_fill()   -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_HEADER_BG)

def _header_align()  -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _alt_fill()      -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_ALT_ROW_BG)

def _green_fill()    -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_GREEN_BG)

def _border() -> Border:
    thin = Side(style="thin", color="D0D5DD")
    return Border(bottom=thin)

def _url_font() -> Font:
    return Font(name="Calibri", color=_LINK_BLUE, underline="single", size=9)

def _data_align(wrap: bool = False) -> Alignment:
    return Alignment(vertical="center", wrap_text=wrap)


# ─────────────────────────────────────────────────────────────────────────────
# LEAD → ROW
# ─────────────────────────────────────────────────────────────────────────────
def _lead_to_row(lead: Dict[str, Any]) -> List[Any]:
    """Convert a lead dict to an ordered list matching SHEET_HEADERS."""
    row: List[Any] = [""] * len(SHEET_HEADERS)
    for field, col in _FIELD_TO_COL.items():
        value = lead.get(field, "")
        row[col - 1] = value if value is not None else ""
    return row


# ─────────────────────────────────────────────────────────────────────────────
# MAIN EXPORT FUNCTION
# ─────────────────────────────────────────────────────────────────────────────
def export_to_excel(
    leads: List[Dict[str, Any]],
    output_path: str = "leads_export.xlsx",
) -> str:
    """
    Write all leads to a formatted Excel workbook.

    Features:
        • Dark navy bold header row with white text, frozen pane
        • Alternating row shading; green highlight for rows with an email
        • Hyperlink formatting on URL, Instagram, TikTok, Twitter, LinkedIn cols
        • Outreach Status dropdown (data validation, no macros)
        • Auto-set column widths
        • Summary sheet with run statistics

    Returns the absolute path to the saved file.
    """
    if not leads:
        raise ValueError("No leads to export.")

    wb = openpyxl.Workbook()

    # ── Main leads sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leads"

    # Write + format header row
    for col_idx, header in enumerate(SHEET_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _header_font()
        cell.fill      = _header_fill()
        cell.alignment = _header_align()
        cell.border    = _border()

    # Freeze the header row so it stays visible while scrolling
    ws.freeze_panes = "A2"

    # Set row height for the header
    ws.row_dimensions[1].height = 28

    # URL column indices (1-based) — these get hyperlink + blue underline
    _url_cols = {2, 7, 8, 9, 10}

    # Write data rows
    for row_num, lead in enumerate(leads, start=2):
        row_data = _lead_to_row(lead)
        has_email = bool(lead.get("email", "").strip())

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = _data_align(wrap=(col_idx == 12))  # wrap description
            cell.font      = Font(name="Calibri", size=9)

            # Alternating row fill; green if email present
            if has_email:
                cell.fill = _green_fill()
            elif row_num % 2 == 0:
                cell.fill = _alt_fill()

            # Hyperlink styling on URL columns
            if col_idx in _url_cols and isinstance(value, str) and value.startswith("http"):
                cell.font      = _url_font()
                cell.hyperlink = value

        # Set a comfortable row height
        ws.row_dimensions[row_num].height = 18

    # Apply column widths
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Outreach Status dropdown (column N = col 14)
    status_formula = f'"{",".join(_STATUS_OPTIONS)}"'
    dv = DataValidation(
        type="list",
        formula1=status_formula,
        allow_blank=True,
        showDropDown=False,   # False = show the arrow button in Excel
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please choose a value from the dropdown list.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"N2:N{len(leads) + 1}"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary")

    total     = len(leads)
    with_email = sum(1 for l in leads if l.get("email", "").strip())
    with_ig    = sum(1 for l in leads if l.get("instagram", "").strip())
    with_tt    = sum(1 for l in leads if l.get("tiktok", "").strip())
    with_tw    = sum(1 for l in leads if l.get("twitter", "").strip())
    keywords   = sorted({l.get("keyword_found_by", "") for l in leads if l.get("keyword_found_by")})
    avg_subs   = (
        int(sum(l.get("subscribers", 0) for l in leads) / total) if total else 0
    )

    summary_rows = [
        ("Metric",                   "Value"),
        ("Export Date",              datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Leads",              total),
        ("Leads with Email",         f"{with_email} ({round(with_email/total*100 if total else 0, 1)}%)"),
        ("Leads with Instagram",     with_ig),
        ("Leads with TikTok",        with_tt),
        ("Leads with Twitter / X",   with_tw),
        ("Average Subscribers",      f"{avg_subs:,}"),
        ("Keywords Searched",        ", ".join(keywords)),
    ]

    for r_idx, (label, value) in enumerate(summary_rows, start=1):
        lc = ws_sum.cell(row=r_idx, column=1, value=label)
        vc = ws_sum.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            lc.font = Font(bold=True, color=_HEADER_FG, name="Calibri", size=10)
            vc.font = Font(bold=True, color=_HEADER_FG, name="Calibri", size=10)
            lc.fill = _header_fill()
            vc.fill = _header_fill()
        else:
            lc.font = Font(bold=True, name="Calibri", size=10)
            vc.font = Font(name="Calibri", size=10)
        lc.alignment = _data_align()
        vc.alignment = _data_align()

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 40

    # Save
    wb.save(output_path)
    return str(Path(output_path).resolve())


# ─────────────────────────────────────────────────────────────────────────────
# GOOGLE SHEETS STUB  (kept so old imports don't break)
# Replace this with the real implementation once you have a Google Cloud account.
# ─────────────────────────────────────────────────────────────────────────────
def export_to_google_sheets(
    leads: List[Dict[str, Any]],
    sheet_name: str,
    credentials_path: str,
) -> str:
    """
    Placeholder — Google Sheets export requires a service account credentials.json.
    Falls back to Excel export and raises a clear message to the UI.
    """
    raise NotImplementedError(
        "Google Sheets export requires a Google Cloud service account. "
        "Use 'Export to Excel' instead — it produces the same formatted output "
        "and can be uploaded to Google Sheets manually for free."
    )
